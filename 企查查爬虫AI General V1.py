import csv
import logging
import random
import time
import re

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from webdriver_manager.chrome import ChromeDriverManager
import os
from bs4 import BeautifulSoup as bs, BeautifulSoup # type: ignore
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import requests
import json

API_KEY = "sk-sqoifenozklojkfvovkvlphfrxrudrfstfqrnrwwrdhwatgi"

# 文件名timestamp (e.g., "20230717_142530" for July 17, 2023, 14:25:30)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")


url = 'https://www.qcc.com/'

# 结果文件路径(自动找desktop)
desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')

result_file_path = os.path.join(desktop_path, f'company_data_{timestamp}.csv')
# e.g result_file_path = os.path.join(desktop_path, 'data0717_01.csv')

# 参数为搜索关键词，调用时传入即可
def crawl_company_info(keyword):
    # 确保目录和文件存在
    directory = os.path.dirname(result_file_path)
    if not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    if not os.path.exists(result_file_path):
        with open(result_file_path, 'w', newline='', encoding='utf-8') as f:
            pass

#    fields = [
 #        '企业名称', '注册资本', '企业类型', '人员规模', '营业场所', '地址变更时间',
  #      '地址变更前', '地址变更后', '经营范围', '负责人', '国标行业', '统一社会信用代码',
   #     '工商注册号', '营业期限', '参保人数', '登记机关', '英文名', '成立日期',
    #    '纳税人识别号', '纳税人资质', '核准日期',  '实缴资本', '进出口企业代码',
     #   '法定代表人', '分支机构参保人数',  '组织机构代码', '所属地区', '登记状态'
    #]

    fields = [
         '企业名称', '注册资本', '人员规模', '营业场所','主要营业场所','手机号', '地址变更时间',
        '地址变更前', '地址变更后', '所有地址变更记录', '经营范围', '负责人', '国标行业',
        '营业期限', '参保人数', '登记机关', '英文名', '成立日期',
        '核准日期',  '实缴资本', '法定代表人', '分支机构参保人数', '所属地区', '登记状态',
        '所有变更记录', '注册地址', '主要经营场所'
    ]
    
    df = pd.DataFrame(columns=fields)
    pd.DataFrame(columns=fields).to_csv(result_file_path, index=False, encoding='utf-8')

    options = webdriver.ChromeOptions()
    # options.add_argument("--headless=new")
    # options.add_argument("--start-minimized")
    # driver = webdriver.Chrome(options=options, service=service)
    service = Service(executable_path="chromedriver")
    # Autodownload chromedriver
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(2)


    # 登录账号
    login_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.qccd-btn.qccd-btn-primary.qcc-header-login-btn"))
    )
    login_button.click()

    print("请60s内扫码登录")
    time.sleep(60)
    driver.refresh()


    # 关键词搜索
    try:
        # 输入关键词
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "searchKey"))
        )
        search_box.send_keys(keyword)
        time.sleep(2)

        # 点击搜索
        search_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "span.input-group-btn button"))
        )
        search_button.click()
        time.sleep(1)

        # 翻页遍历
        try:
    # 获取总页数（带错误处理）
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//ul[@class='pagination']/li[last()-1]/a"))
                )
                total_pages = int(re.search(r'\d+', element.text).group())
                print(f"一共 {total_pages} 页！")
            except Exception as e:
                total_pages = 1
                print(f"未找到分页控件，默认处理为1页。错误: {str(e)}")        

            for page in range(1, total_pages + 1):
                print(f"正在遍历第 {page} 页")

                # 遍历公司信息
                try:
                    # 点击公司
                    company_elements = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.title.copy-value"))
                    )
                    for company in company_elements:
                        company_name = company.text
                        print(f"正在爬取公司: {company_name}")

                        # 当前窗口句柄
                        main_window = driver.current_window_handle

                        company.click()
                        time.sleep(random.uniform(0.5, 1.5))

                        # 等待新窗口加载
                        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

                        # 切换到新窗口
                        for window_handle in driver.window_handles:
                            if window_handle != main_window:
                                driver.switch_to.window(window_handle)
                                break

                        # 获取公司基础信息及地址变更信息
                        try:
                            # 基础信息
                            comp_info_parent = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located(
                                    (By.CSS_SELECTOR, "div.cominfo-normal")
                                )
                            )
                            comp_info_element = comp_info_parent.find_element(By.CSS_SELECTOR, "table.ntable")
                            comp_info = comp_info_element.get_attribute('outerHTML')

                            soup1 = BeautifulSoup(comp_info, 'html.parser')
                            table1 = soup1.find('table', class_='ntable')

                            # 将所有基础信息存进列表
                            comp_info_data = []
                            for row in table1.find_all('tr'):
                                row_data = []
                                for cell in row.find_all(['td', 'th']):
                                    row_data.append(cell.get_text(strip=True))
                                comp_info_data.append(row_data)

                            # 地址变更(可能不存在)
                            address_change_data = []
                            try:
                                address_change_element = WebDriverWait(driver, 3).until(
                                    EC.presence_of_element_located(
                                        (By.CSS_SELECTOR, 'table.ntable.app-ntable-expand-all.hide-info'))
                                )
                                address_change = address_change_element.get_attribute('outerHTML')

                                soup2 = BeautifulSoup(address_change, 'html.parser')
                                table2 = soup2.find('table', class_='ntable app-ntable-expand-all hide-info')

                                # 将所有地址变更信息存进列表
                                for row in table2.find_all('tr'):
                                    row_data = []
                                    for cell in row.find_all(['td']):
                                        row_data.append(cell.get_text(strip=True))
                                    address_change_data.append(row_data)

                            except Exception as e:
                                print(f"公司无变更信息！")

                            # 存放最终result
                            row_data = {field: "" for field in fields}

                            # 将基础信息结构化
                            for row in comp_info_data:
                                # 确保行数据至少包含2个元素（字段名和值）
                                if len(row) < 2:
                                    continue
                                # 偶数为字段名，奇数为值
                                for i in range(0, len(row), 2):
                                    # 确保值的索引存在
                                    if i + 1 < len(row):
                                        field_name = row[i].replace(' ', '')
                                        field_value = row[i + 1]

                                        # 仅处理目标字段列表中的字段，且值为空时才覆盖
                                        if field_name in fields and not row_data[field_name]:
                                            row_data[field_name] = field_value

                            # 将地址变更信息结构化
                            full_address_changes = []
                            for row in address_change_data:
                                if len(row) >= 4 and '地址' in row[2]:
                                    change_time = row[1]
                                    address_old = row[3]
                                    address_new = row[4]

                                    if '地址变更时间' not in row_data:  # 仅首次赋值（最新记录在前）
                                        row_data['地址变更时间'] = change_time
                                        row_data['地址变更前'] = address_old
                                        row_data['地址变更后'] = address_new

                                        # 2. 收集所有变更记录，格式："时间|变更前|变更后"
                                    full_address_changes.append(f"{change_time}|{address_old}|{address_new}")

                                    # 3. 新增字段：所有地址变更记录（用分号分隔）
                                    row_data['所有地址变更记录'] = ';'.join(full_address_changes)

                            # 将地址变更信息结构化
                            full_changes = []
                            for row in address_change_data:
                                if len(row) >= 4:
                                    change_time = row[1]
                                    change_old = row[3]
                                    change_new = row[4]

                                    full_changes.append(f"{change_time}|{change_old}|{change_new}")

                                    # 3. 新增字段：所有地址变更记录（用分号分隔）
                                    row_data['所有变更记录'] = ';'.join(full_changes)

                            # 手机号
                            phone_number = None
                            try:

                                phone_element = WebDriverWait(driver, 5).until(
                                    EC.presence_of_element_located(
                                        (By.CSS_SELECTOR,
                                         "span.f span.app-copy-box span.val span.copy-value.need-copy-field")
                                    )
                                )

                                phone_number = phone_element.get_attribute('textContent').strip()
                                row_data['手机号'] = phone_number

                            except Exception as e:
                                row_data['手机号'] = ""
                                print(f"未找到手机号信息：{str(e)}")


                            # 写入文件

                            if str(row_data.get('登记状态', '')).startswith('存续'):
                                pd.DataFrame([row_data]).to_csv(
                                    result_file_path,
                                    mode='a',
                                    header=False,
                                    index=False,
                                    encoding='utf-8'
                                )
                            else:
                                print(f"跳过: {row_data.get('企业名称')} (状态: {row_data.get('登记状态')})")
                            time.sleep(random.uniform(0.5, 1.5))

                        except Exception as e:
                            print(f"公司信息爬取错误：{str(e)}")
                            continue

                        finally:
                            # 关闭新窗口并切回主窗口
                            if driver.current_window_handle != main_window:
                                driver.close()
                            driver.switch_to.window(main_window)

                            time.sleep(random.uniform(1, 1.5))

                except Exception as e:
                    print(f"公司遍历错误：{str(e)}")
                    break

                # 如果不是最后一页，点击翻页
                if page < total_pages:
                    next_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//a[contains(text(), ">")]'))
                    )
                    next_button.click()
                    time.sleep(2)

        except Exception as e:
            print(f"翻页遍历错误：{str(e)}")

    except Exception as e:
        print(f"关键词搜索错误：{str(e)}")

    finally:
        driver.quit()


def csv_to_excel_with_highlight(csv_file_path, excel_file_path):
    try:
        # 读取 CSV 文件（可按需添加编码参数，如 encoding='utf-8'）
        df = pd.read_csv(csv_file_path)

        # 保存为 Excel 文件
        df.to_excel(excel_file_path, index=False, engine='openpyxl')

        # 加载 Excel 文件进行格式设置
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # 定义颜色填充
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # 获取列名所在的行号（通常是第一行）
        header_row = 1

        # 获取各列的列号，增加默认值避免后续报错
        capital_col = None
        employee_col = None
        address_change_col = None

        # 查找列名对应的列号（支持模糊匹配，如“注册资本”“参保人数”“地址变更后”）
        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=header_row, column=col).value
            if col_name and re.search("注册资本", col_name):
                capital_col = col
            elif col_name and re.search("参保人数", col_name):
                employee_col = col
            elif col_name and re.search("地址变更后", col_name):
                address_change_col = col

        # 遍历每一行数据，从 header_row + 1 开始（跳过表头）
        for row in range(header_row + 1, ws.max_row + 1):
            # 标红逻辑：仅当“地址变更后”列存在时执行
            if address_change_col is not None:
                address_val = ws.cell(row=row, column=address_change_col).value
                # 清洗值：去除空白、判断非空
                clean_addr = str(address_val).strip() if pd.notna(address_val) else ""
                if clean_addr != "":
                    # 标红整行
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = green_fill
                    # 标红后跳过标黄逻辑
                    continue

            # 标黄逻辑：需“注册资本”和“参保人数”列存在
            if capital_col is not None and employee_col is not None:
                # 处理注册资本
                capital_val = ws.cell(row=row, column=capital_col).value
                capital_over = False
                if pd.notna(capital_val):
                    # 清洗逻辑：提取“万元”数值（假设仅处理人民币万元，可扩展）
                    cap_match = re.search(r"(\d+(\.\d+)?)万元", str(capital_val))
                    if cap_match:
                        capital_num = float(cap_match.group(1)) * 10000  # 转换为元
                        capital_over = capital_num > capitalreq * 10000  # 判断是否超500万（元）

                # 处理参保人数
                employee_val = ws.cell(row=row, column=employee_col).value
                employee_over = False
                if pd.notna(employee_val):
                    # 清洗逻辑：提取数字（如处理“5(2024年报)”“<null>”等）
                    emp_match = re.search(r"(\d+)", str(employee_val))
                    if emp_match:
                        employee_num = int(emp_match.group(1))
                        employee_over = employee_num > employeereq

                # 标黄条件：注册资本超500万 或 人数超30
                if capital_over or employee_over:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill

        # 保存修改后的 Excel 文件
        wb.save(excel_file_path)
        print(f"成功转换并标记颜色，文件已保存至: {excel_file_path}")

    except Exception as e:
        print(f"转换过程中出错: {str(e)}")


def excel_to_json_string(excel_path):
    """Convert Excel to JSON string variable instead of file"""
    try:
        df = pd.read_excel(excel_path, dtype=str)
        df = df.where(pd.notnull(df), None)
        
        # Convert to list of dictionaries
        records = []
        for record in df.to_dict(orient='records'):
            # Clean the record - remove None and empty values
            clean_record = {}
            for k, v in record.items():
                if v is not None and str(v).strip() not in ['', '-', 'nan', 'None']:
                    clean_record[str(k)] = str(v).strip()
            if clean_record:  # Only add non-empty records
                records.append(clean_record)
        
        # Convert to JSON string
        json_string = json.dumps(records, ensure_ascii=False, indent=2)
        
        print("✅ Successfully converted Excel to JSON string")
        print(f"📊 Total records: {len(records)}")
        print("\n📋 First 3 records in JSON format:")
        print(json.dumps(records[:3], ensure_ascii=False, indent=2))
        
        return json_string
        
    except Exception as e:
        print(f"❌ Conversion failed: {str(e)}")
        return None

def deepseek_text_analysis(json_string):
    """Send JSON string data to DeepSeek API"""
    try:
        url = "https://api.siliconflow.cn/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "deepseek-ai/DeepSeek-V3",
            "messages": [{
                "role": "user",
                "content": f"""请分析以下企业数据，作为一个戴德梁行房地产顾问，分析哪些公司适合去B（可以推荐10个或者你觉得合适的数量），给出原因并且列出公司基本信息+电话，去bd的流程 如要联系谁，请考虑公司所在行业发展前景，公司变更信息分析有没有可能搬迁，如可能租约快要到期，或者近期资金增加或者缩小规模导致可能搬迁，以及其他你觉得可能重要的因素，请细致的分析，最后请根据这你推荐的这几个公司制作一个拜访的先后顺序和地图上的最高效率路线顺序。回答请正式一点，不需要说无关话语。

企业数据（JSON格式）：
{json_string}

请按以下格式分析：
1. 公司名称和基本信息
2. 推荐理由（基于注册资本、参保人数、地址变更等）
3. BD优先级排序（1-5星）
4. 建议的BD路线"""
            }],
            "temperature": 0.7,
            "max_tokens": 2000
        }
        
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code == 200:
            return response.json()['choices'][0]['message']['content']
        else:
            print(f"Analysis failed: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Analysis error: {str(e)}")
        return None

def deepseekinput(excel_file_path):
    """Convert Excel to JSON string and analyze with DeepSeek"""
    print("正在使用DeepSeek AI分析数据...")
    
    # Convert Excel to JSON string
    json_string = excel_to_json_string(excel_file_path)
    
    if not json_string:
        print("❌ Failed to convert Excel to JSON")
        return None
    
    # Send JSON string to DeepSeek for analysis
    result = deepseek_text_analysis(json_string)
    
    return result


if __name__ == '__main__':
    #输入搜索关键词
    userinput=input("请输入搜索关键词: ").strip()
    capitalreq=int(input("注册资金大于标准（万）标黄："))
    employeereq=int(input("员工数量大于（人数）标黄："))
    print("-----------------")
    print("公司地址近期有变更为绿色高亮")
    crawl_company_info(userinput)

    excel_file_path = os.path.join(desktop_path, f'company_data_{timestamp}.xlsx')
    csv_to_excel_with_highlight(
        csv_file_path=result_file_path,
        excel_file_path=excel_file_path
    )
    
    # 调用DeepSeek API进行数据分析
    print("正在使用DeepSeek AI分析数据...")
    
    # Convert Excel to JSON string and analyze
    json_string = excel_to_json_string(excel_file_path)
    
    if json_string:
        # Send JSON string to DeepSeek for analysis
        analysis_result = deepseek_text_analysis(json_string)
        
        if analysis_result:
            # 将分析结果添加到Excel文件的单独工作表中
            try:
                # 加载现有的Excel文件
                wb = openpyxl.load_workbook(excel_file_path)
                
                # 创建新的工作表用于分析结果
                if 'AI分析结果' in wb.sheetnames:
                    ws_analysis = wb['AI分析结果']
                    ws_analysis.delete_rows(1, ws_analysis.max_row)  # 清空现有内容
                else:
                    ws_analysis = wb.create_sheet('AI分析结果')
                
                # 添加标题
                ws_analysis['A1'] = 'DeepSeek AI 分析结果'
                ws_analysis['A1'].font = openpyxl.styles.Font(bold=True, size=14)
                
                # 添加分隔线
                ws_analysis['A2'] = '=' * 50
                
                # 将分析结果按行分割并写入
                analysis_lines = analysis_result.split('\n')
                for i, line in enumerate(analysis_lines, start=3):
                    ws_analysis[f'A{i}'] = line
                
                # 调整列宽
                ws_analysis.column_dimensions['A'].width = 100
                
                # 保存Excel文件
                wb.save(excel_file_path)
                print(f"分析完成！结果已添加到Excel文件的工作表: AI分析结果")
                print("\n分析结果:")
                print("=" * 50)
                print(analysis_result)
                
            except Exception as e:
                print(f"保存分析结果到Excel时出错: {str(e)}")
                # 如果保存到Excel失败，仍然保存为文本文件作为备份
                analysis_file_path = os.path.join(desktop_path, f'analysis_result_{timestamp}.txt')
                with open(analysis_file_path, 'w', encoding='utf-8') as f:
                    f.write("DeepSeek AI 分析结果\n")
                    f.write("=" * 50 + "\n")
                    f.write(analysis_result)
                print(f"已保存备份文件: {analysis_file_path}")
        else:
            print("分析失败，请检查API或网络连接")
    else:
        print("Excel转JSON失败，无法进行分析")
