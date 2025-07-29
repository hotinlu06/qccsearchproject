import csv
import logging
import random
import time
import re

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import os
from bs4 import BeautifulSoup as bs, BeautifulSoup # type: ignore
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# æ–‡ä»¶åtimestamp (e.g., "20230717_142530" for July 17, 2023, 14:25:30)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")


url = 'https://www.qcc.com/'

# ç»“æœæ–‡ä»¶è·¯å¾„(è‡ªåŠ¨æ‰¾desktop)
desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')

result_file_path = os.path.join(desktop_path, f'company_data_{timestamp}.csv')
# e.g result_file_path = os.path.join(desktop_path, 'data0717_01.csv')

# å‚æ•°ä¸ºæœç´¢å…³é”®è¯ï¼Œè°ƒç”¨æ—¶ä¼ å…¥å³å¯
def crawl_company_info(keyword):
    # ç¡®ä¿ç›®å½•å’Œæ–‡ä»¶å­˜åœ¨
    directory = os.path.dirname(result_file_path)
    if not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    if not os.path.exists(result_file_path):
        with open(result_file_path, 'w', newline='', encoding='utf-8') as f:
            pass

#    fields = [
 #        'ä¼ä¸šåç§°', 'æ³¨å†Œèµ„æœ¬', 'ä¼ä¸šç±»å‹', 'äººå‘˜è§„æ¨¡', 'è¥ä¸šåœºæ‰€', 'åœ°å€å˜æ›´æ—¶é—´',
  #      'åœ°å€å˜æ›´å‰', 'åœ°å€å˜æ›´å', 'ç»è¥èŒƒå›´', 'è´Ÿè´£äºº', 'å›½æ ‡è¡Œä¸š', 'ç»Ÿä¸€ç¤¾ä¼šä¿¡ç”¨ä»£ç ',
   #     'å·¥å•†æ³¨å†Œå·', 'è¥ä¸šæœŸé™', 'å‚ä¿äººæ•°', 'ç™»è®°æœºå…³', 'è‹±æ–‡å', 'æˆç«‹æ—¥æœŸ',
    #    'çº³ç¨äººè¯†åˆ«å·', 'çº³ç¨äººèµ„è´¨', 'æ ¸å‡†æ—¥æœŸ',  'å®ç¼´èµ„æœ¬', 'è¿›å‡ºå£ä¼ä¸šä»£ç ',
     #   'æ³•å®šä»£è¡¨äºº', 'åˆ†æ”¯æœºæ„å‚ä¿äººæ•°',  'ç»„ç»‡æœºæ„ä»£ç ', 'æ‰€å±åœ°åŒº', 'ç™»è®°çŠ¶æ€'
    #]

    fields = [
         'ä¼ä¸šåç§°', 'æ³¨å†Œèµ„æœ¬', 'äººå‘˜è§„æ¨¡', 'è¥ä¸šåœºæ‰€', 'åœ°å€å˜æ›´æ—¶é—´',
        'åœ°å€å˜æ›´å‰', 'åœ°å€å˜æ›´å', 'ç»è¥èŒƒå›´', 'è´Ÿè´£äºº', 'å›½æ ‡è¡Œä¸š',
        'è¥ä¸šæœŸé™', 'å‚ä¿äººæ•°', 'ç™»è®°æœºå…³', 'è‹±æ–‡å', 'æˆç«‹æ—¥æœŸ',
        'æ ¸å‡†æ—¥æœŸ',  'å®ç¼´èµ„æœ¬', 'æ³•å®šä»£è¡¨äºº', 'åˆ†æ”¯æœºæ„å‚ä¿äººæ•°', 'æ‰€å±åœ°åŒº', 'ç™»è®°çŠ¶æ€'
    ]
    
    df = pd.DataFrame(columns=fields)
    pd.DataFrame(columns=fields).to_csv(result_file_path, index=False, encoding='utf-8')

    options = webdriver.ChromeOptions()
    # options.add_argument("--headless=new")
    # options.add_argument("--start-minimized")
    # driver = webdriver.Chrome(options=options, service=service)
    # service = Service(executable_path="chromedriver") (manual installation for chromedriver)
    # Autodownload chromedriver

##windows
        
   # options = webdriver.ChromeOptions()
    #options.add_argument('--ignore-certificate-errors')  # ğŸ‘ˆ Critical for Windows
    #options.add_argument('--allow-running-insecure-content')
    #options.add_argument('--disable-extensions')

# Add these for corporate networks
    options.add_argument('--proxy-server="direct://"')
    options.add_argument('--proxy-bypass-list=*')

    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(2)

    # ç™»å½•è´¦å·
    login_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.qccd-btn.qccd-btn-primary.qcc-header-login-btn"))
    )
    login_button.click()

    print("è¯·60så†…æ‰«ç ç™»å½•")
    time.sleep(60)
    driver.refresh()




    # å…³é”®è¯æœç´¢
    try:
        # è¾“å…¥å…³é”®è¯
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "searchKey"))
        )
        search_box.send_keys(keyword)
        time.sleep(2)

        # ç‚¹å‡»æœç´¢
        search_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "span.input-group-btn button"))
        )
        search_button.click()
        time.sleep(1)

        # ç¿»é¡µéå†
        try:
    # è·å–æ€»é¡µæ•°ï¼ˆå¸¦é”™è¯¯å¤„ç†ï¼‰
            try:
                element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//ul[@class='pagination']/li[last()-1]/a"))
                )
                total_pages = int(re.search(r'\d+', element.text).group())
                print(f"ä¸€å…± {total_pages} é¡µï¼")
            except Exception as e:
                total_pages = 1
                print(f"æœªæ‰¾åˆ°åˆ†é¡µæ§ä»¶ï¼Œé»˜è®¤å¤„ç†ä¸º1é¡µã€‚é”™è¯¯: {str(e)}")        

            for page in range(1, total_pages + 1):
                print(f"æ­£åœ¨éå†ç¬¬ {page} é¡µ")

                # éå†å…¬å¸ä¿¡æ¯
                try:
                    # ç‚¹å‡»å…¬å¸
                    company_elements = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.title.copy-value"))
                    )
                    for company in company_elements:
                        company_name = company.text
                        print(f"æ­£åœ¨çˆ¬å–å…¬å¸: {company_name}")

                        # å½“å‰çª—å£å¥æŸ„
                        main_window = driver.current_window_handle

                        company.click()
                        time.sleep(random.uniform(0.5, 1.5))

                        # ç­‰å¾…æ–°çª—å£åŠ è½½
                        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

                        # åˆ‡æ¢åˆ°æ–°çª—å£
                        for window_handle in driver.window_handles:
                            if window_handle != main_window:
                                driver.switch_to.window(window_handle)
                                break

                        # è·å–å…¬å¸åŸºç¡€ä¿¡æ¯åŠåœ°å€å˜æ›´ä¿¡æ¯
                        try:
                            # åŸºç¡€ä¿¡æ¯
                            comp_info_parent = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located(
                                    (By.CSS_SELECTOR, "div.cominfo-normal")
                                )
                            )
                            comp_info_element = comp_info_parent.find_element(By.CSS_SELECTOR, "table.ntable")
                            comp_info = comp_info_element.get_attribute('outerHTML')

                            soup1 = BeautifulSoup(comp_info, 'html.parser')
                            table1 = soup1.find('table', class_='ntable')

                            # å°†æ‰€æœ‰åŸºç¡€ä¿¡æ¯å­˜è¿›åˆ—è¡¨
                            comp_info_data = []
                            for row in table1.find_all('tr'):
                                row_data = []
                                for cell in row.find_all(['td', 'th']):
                                    row_data.append(cell.get_text(strip=True))
                                comp_info_data.append(row_data)

                            # åœ°å€å˜æ›´(å¯èƒ½ä¸å­˜åœ¨)
                            address_change_data = []
                            try:
                                address_change_element = WebDriverWait(driver, 3).until(
                                    EC.presence_of_element_located(
                                        (By.CSS_SELECTOR, 'table.ntable.app-ntable-expand-all.hide-info'))
                                )
                                address_change = address_change_element.get_attribute('outerHTML')

                                soup2 = BeautifulSoup(address_change, 'html.parser')
                                table2 = soup2.find('table', class_='ntable app-ntable-expand-all hide-info')

                                # å°†æ‰€æœ‰åœ°å€å˜æ›´ä¿¡æ¯å­˜è¿›åˆ—è¡¨
                                for row in table2.find_all('tr'):
                                    row_data = []
                                    for cell in row.find_all(['td']):
                                        row_data.append(cell.get_text(strip=True))
                                    address_change_data.append(row_data)

                            except Exception as e:
                                print(f"å…¬å¸æ— å˜æ›´ä¿¡æ¯ï¼")

                            # å­˜æ”¾æœ€ç»ˆresult
                            row_data = {field: "" for field in fields}

                            # å°†åŸºç¡€ä¿¡æ¯ç»“æ„åŒ–
                            for row in comp_info_data:
                                # ç¡®ä¿è¡Œæ•°æ®è‡³å°‘åŒ…å«2ä¸ªå…ƒç´ ï¼ˆå­—æ®µåå’Œå€¼ï¼‰
                                if len(row) < 2:
                                    continue
                                # å¶æ•°ä¸ºå­—æ®µåï¼Œå¥‡æ•°ä¸ºå€¼
                                for i in range(0, len(row), 2):
                                    # ç¡®ä¿å€¼çš„ç´¢å¼•å­˜åœ¨
                                    if i + 1 < len(row):
                                        field_name = row[i].replace(' ', '')
                                        field_value = row[i + 1]

                                        # ä»…å¤„ç†ç›®æ ‡å­—æ®µåˆ—è¡¨ä¸­çš„å­—æ®µï¼Œä¸”å€¼ä¸ºç©ºæ—¶æ‰è¦†ç›–
                                        if field_name in fields and not row_data[field_name]:
                                            row_data[field_name] = field_value

                            # å°†åœ°å€å˜æ›´ä¿¡æ¯ç»“æ„åŒ–
                            for row in address_change_data:
                                if len(row) >= 4 and 'åœ°å€' in row[2]:
                                    change_time = row[1]
                                    address_old = row[3]
                                    address_new = row[4]

                                    # æœ€æ–°çš„æ•°æ®æ’åœ¨è¶Šå‰é¢ï¼Œåªéœ€éå†ç¬¬ä¸€ä¸ªå³å¯
                                    row_data['åœ°å€å˜æ›´æ—¶é—´'] = change_time
                                    row_data['åœ°å€å˜æ›´å‰'] = address_old
                                    row_data['åœ°å€å˜æ›´å'] = address_new
                                    break

                            # å†™å…¥æ–‡ä»¶

                            if str(row_data.get('ç™»è®°çŠ¶æ€', '')).startswith('å­˜ç»­'):
                                pd.DataFrame([row_data]).to_csv(
                                    result_file_path,
                                    mode='a',
                                    header=False,
                                    index=False,
                                    encoding='utf-8'
                                )
                            else:
                                print(f"è·³è¿‡: {row_data.get('ä¼ä¸šåç§°')} (çŠ¶æ€: {row_data.get('ç™»è®°çŠ¶æ€')})")
                            time.sleep(random.uniform(0.5, 1.5))

                        except Exception as e:
                            print(f"å…¬å¸ä¿¡æ¯çˆ¬å–é”™è¯¯ï¼š{str(e)}")
                            continue

                        finally:
                            # å…³é—­æ–°çª—å£å¹¶åˆ‡å›ä¸»çª—å£
                            if driver.current_window_handle != main_window:
                                driver.close()
                            driver.switch_to.window(main_window)

                            time.sleep(random.uniform(1, 1.5))

                except Exception as e:
                    print(f"å…¬å¸éå†é”™è¯¯ï¼š{str(e)}")
                    break

                # å¦‚æœä¸æ˜¯æœ€åä¸€é¡µï¼Œç‚¹å‡»ç¿»é¡µ
                if page < total_pages:
                    next_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//a[contains(text(), ">")]'))
                    )
                    next_button.click()
                    time.sleep(2)

        except Exception as e:
            print(f"ç¿»é¡µéå†é”™è¯¯ï¼š{str(e)}")

    except Exception as e:
        print(f"å…³é”®è¯æœç´¢é”™è¯¯ï¼š{str(e)}")

    finally:
        driver.quit()


def csv_to_excel_with_highlight(csv_file_path, excel_file_path):
    try:
        # è¯»å– CSV æ–‡ä»¶ï¼ˆå¯æŒ‰éœ€æ·»åŠ ç¼–ç å‚æ•°ï¼Œå¦‚ encoding='utf-8'ï¼‰
        df = pd.read_csv(csv_file_path)

        # ä¿å­˜ä¸º Excel æ–‡ä»¶
        df.to_excel(excel_file_path, index=False, engine='openpyxl')

        # åŠ è½½ Excel æ–‡ä»¶è¿›è¡Œæ ¼å¼è®¾ç½®
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # å®šä¹‰é¢œè‰²å¡«å……
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # è·å–åˆ—åæ‰€åœ¨çš„è¡Œå·ï¼ˆé€šå¸¸æ˜¯ç¬¬ä¸€è¡Œï¼‰
        header_row = 1

        # è·å–å„åˆ—çš„åˆ—å·ï¼Œå¢åŠ é»˜è®¤å€¼é¿å…åç»­æŠ¥é”™
        capital_col = None
        employee_col = None
        address_change_col = None

        # æŸ¥æ‰¾åˆ—åå¯¹åº”çš„åˆ—å·ï¼ˆæ”¯æŒæ¨¡ç³ŠåŒ¹é…ï¼Œå¦‚â€œæ³¨å†Œèµ„æœ¬â€â€œå‚ä¿äººæ•°â€â€œåœ°å€å˜æ›´åâ€ï¼‰
        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=header_row, column=col).value
            if col_name and re.search("æ³¨å†Œèµ„æœ¬", col_name):
                capital_col = col
            elif col_name and re.search("å‚ä¿äººæ•°", col_name):
                employee_col = col
            elif col_name and re.search("åœ°å€å˜æ›´å", col_name):
                address_change_col = col

        # éå†æ¯ä¸€è¡Œæ•°æ®ï¼Œä» header_row + 1 å¼€å§‹ï¼ˆè·³è¿‡è¡¨å¤´ï¼‰
        for row in range(header_row + 1, ws.max_row + 1):
            # æ ‡çº¢é€»è¾‘ï¼šä»…å½“â€œåœ°å€å˜æ›´åâ€åˆ—å­˜åœ¨æ—¶æ‰§è¡Œ
            if address_change_col is not None:
                address_val = ws.cell(row=row, column=address_change_col).value
                # æ¸…æ´—å€¼ï¼šå»é™¤ç©ºç™½ã€åˆ¤æ–­éç©º
                clean_addr = str(address_val).strip() if pd.notna(address_val) else ""
                if clean_addr != "":
                    # æ ‡çº¢æ•´è¡Œ
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = green_fill
                    # æ ‡çº¢åè·³è¿‡æ ‡é»„é€»è¾‘
                    continue

            # æ ‡é»„é€»è¾‘ï¼šéœ€â€œæ³¨å†Œèµ„æœ¬â€å’Œâ€œå‚ä¿äººæ•°â€åˆ—å­˜åœ¨
            if capital_col is not None and employee_col is not None:
                # å¤„ç†æ³¨å†Œèµ„æœ¬
                capital_val = ws.cell(row=row, column=capital_col).value
                capital_over = False
                if pd.notna(capital_val):
                    # æ¸…æ´—é€»è¾‘ï¼šæå–â€œä¸‡å…ƒâ€æ•°å€¼ï¼ˆå‡è®¾ä»…å¤„ç†äººæ°‘å¸ä¸‡å…ƒï¼Œå¯æ‰©å±•ï¼‰
                    cap_match = re.search(r"(\d+(\.\d+)?)ä¸‡å…ƒ", str(capital_val))
                    if cap_match:
                        capital_num = float(cap_match.group(1)) * 10000  # è½¬æ¢ä¸ºå…ƒ
                        capital_over = capital_num > capitalreq * 10000  # åˆ¤æ–­æ˜¯å¦è¶…500ä¸‡ï¼ˆå…ƒï¼‰

                # å¤„ç†å‚ä¿äººæ•°
                employee_val = ws.cell(row=row, column=employee_col).value
                employee_over = False
                if pd.notna(employee_val):
                    # æ¸…æ´—é€»è¾‘ï¼šæå–æ•°å­—ï¼ˆå¦‚å¤„ç†â€œ5(2024å¹´æŠ¥)â€â€œ<null>â€ç­‰ï¼‰
                    emp_match = re.search(r"(\d+)", str(employee_val))
                    if emp_match:
                        employee_num = int(emp_match.group(1))
                        employee_over = employee_num > employeereq

                # æ ‡é»„æ¡ä»¶ï¼šæ³¨å†Œèµ„æœ¬è¶…500ä¸‡ æˆ– äººæ•°è¶…30
                if capital_over or employee_over:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill

        # ä¿å­˜ä¿®æ”¹åçš„ Excel æ–‡ä»¶
        wb.save(excel_file_path)
        print(f"æˆåŠŸè½¬æ¢å¹¶æ ‡è®°é¢œè‰²ï¼Œæ–‡ä»¶å·²ä¿å­˜è‡³: {excel_file_path}")

    except Exception as e:
        print(f"è½¬æ¢è¿‡ç¨‹ä¸­å‡ºé”™: {str(e)}")



if __name__ == '__main__':
    #è¾“å…¥æœç´¢å…³é”®è¯
    userinput=input("è¯·è¾“å…¥æœç´¢å…³é”®è¯: ").strip()
    capitalreq=int(input("æ³¨å†Œèµ„é‡‘å¤§äºæ ‡å‡†ï¼ˆä¸‡ï¼‰æ ‡é»„ï¼š"))
    employeereq=int(input("å‘˜å·¥æ•°é‡å¤§äºï¼ˆäººæ•°ï¼‰æ ‡é»„ï¼š"))
    print("-----------------")
    print("å…¬å¸åœ°å€è¿‘æœŸæœ‰å˜æ›´ä¸ºç»¿è‰²é«˜äº®")
    crawl_company_info(userinput)

    csv_to_excel_with_highlight(
        csv_file_path=result_file_path,
        excel_file_path = os.path.join(desktop_path, f'company_data_{timestamp}.xlsx')
    )
